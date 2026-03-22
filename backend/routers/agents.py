"""
External agents (sub-agents / agencies / partners) — NOT internal staff.
These are the external entities students belong to.
"""
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session

from backend.database import get_db
from backend.auth import get_current_user
import backend.models as models
import backend.schemas as schemas

router = APIRouter(prefix="/agents", tags=["agents"])


@router.get("/", response_model=List[schemas.AgentOut])
def list_agents(
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: models.User = Depends(get_current_user),
):
    q = db.query(models.Agent)
    if search:
        q = q.filter(models.Agent.name.ilike(f"%{search}%"))
    agents = q.order_by(models.Agent.name).all()
    result = []
    for agent in agents:
        d = schemas.AgentOut.model_validate(agent)
        mapping = agent.manager_mappings[0] if agent.manager_mappings else None
        d.manager_id = mapping.manager_id if mapping else None
        result.append(d)
    return result


@router.post("/", response_model=schemas.AgentOut)
def create_agent(
    data: schemas.AgentCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Admin or manager required")
    agent = models.Agent(**data.model_dump())
    db.add(agent)
    db.commit()
    db.refresh(agent)
    return agent


@router.put("/{agent_id}", response_model=schemas.AgentOut)
def update_agent(
    agent_id: int,
    data: schemas.AgentUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Admin or manager required")
    agent = db.query(models.Agent).filter(models.Agent.id == agent_id).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(agent, k, v)
    db.commit()
    db.refresh(agent)
    return agent


@router.delete("/{agent_id}")
def delete_agent(
    agent_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role not in ("admin",):
        raise HTTPException(status_code=403, detail="Admin only")
    agent = db.query(models.Agent).filter(models.Agent.id == agent_id).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    db.delete(agent)
    db.commit()
    return {"message": "Agent deleted"}


# ─── Bulk Upload ───────────────────────────────────────────────────────────────

@router.post("/bulk-upload", response_model=schemas.BulkUploadResult)
async def bulk_upload_agents(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Upload agents from Excel (.xlsx). Required columns: Agent Name.
    Optional columns: Company, Email, Phone, Country, Manager Name.
    Manager Name maps to an existing user with role manager/admin.
    """
    if current_user.role not in ("admin", "manager"):
        # Also allow custom-role users who have view access to the agents module
        from backend.models import RolePermission
        perm = db.query(RolePermission).filter(
            RolePermission.role == current_user.role,
            RolePermission.department == "agents",
            RolePermission.can_view == True,
        ).first()
        if not perm:
            raise HTTPException(status_code=403, detail="Admin or manager required")

    import openpyxl
    import io

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Excel file. Please upload a .xlsx file.")

    # Read header row
    headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row, name):
        aliases = {
            "agent name": ["agent name", "name", "agent"],
            "company": ["company", "company / agency", "company/agency", "agency"],
            "email": ["email"],
            "phone": ["phone", "telephone", "mobile"],
            "country": ["country"],
            "manager name": ["manager name", "manager", "assigned manager", "agent manager", "agent manager name"],
        }
        for alias in aliases.get(name, []):
            if alias in headers:
                idx = headers.index(alias)
                cell = row[idx]
                return str(cell.value or "").strip() if cell.value is not None else ""
        return ""

    # Build manager name lookup — admin, manager, team_leader +
    # custom roles with can_view for GS or Offer Applications
    from backend.models import RolePermission as _RP
    from sqlalchemy import or_ as _or
    custom_role_names = {
        r.role for r in db.query(_RP.role).filter(
            _RP.can_view == True,
            _or(_RP.department == "gs", _RP.department == "offer"),
        ).distinct().all()
    }
    all_eligible_roles = {"admin", "manager", "team_leader"} | custom_role_names
    managers = db.query(models.User).filter(
        models.User.role.in_(list(all_eligible_roles)),
        models.User.is_active == True,
    ).all()
    manager_map = {m.full_name.lower(): m for m in managers}

    created = 0
    updated = 0
    skipped = 0
    errors = []

    rows = list(ws.iter_rows(min_row=2))
    for i, row in enumerate(rows, start=2):
        name = col(row, "agent name")
        if not name:
            skipped += 1
            continue

        company = col(row, "company") or None
        email = col(row, "email") or None
        phone = col(row, "phone") or None
        country = col(row, "country") or None
        manager_name = col(row, "manager name")

        # Check if agent already exists by name
        existing = db.query(models.Agent).filter(
            models.Agent.name.ilike(name)
        ).first()

        if existing:
            # Agent exists — update fields and always sync manager from sheet
            field_changed = False
            if company and existing.company_name != company:
                existing.company_name = company
                field_changed = True
            if email and existing.email != email:
                existing.email = email
                field_changed = True
            if phone and existing.phone != phone:
                existing.phone = phone
                field_changed = True
            if country and existing.country != country:
                existing.country = country
                field_changed = True

            manager_changed = False
            if manager_name:
                manager = manager_map.get(manager_name.lower())
                if manager:
                    existing_mapping = db.query(models.ManagerAgentMapping).filter_by(
                        agent_id=existing.id
                    ).first()
                    if existing_mapping:
                        if existing_mapping.manager_id != manager.id:
                            # Different manager — update the mapping
                            existing_mapping.manager_id = manager.id
                            manager_changed = True
                        # else: same manager already — no change needed
                    else:
                        # No manager yet — assign
                        db.add(models.ManagerAgentMapping(manager_id=manager.id, agent_id=existing.id))
                        manager_changed = True
                else:
                    errors.append(f"Row {i}: '{name}' exists but manager '{manager_name}' not found — no change.")

            if field_changed or manager_changed:
                updated += 1
            else:
                skipped += 1
            continue

        # New agent — create it
        agent = models.Agent(
            name=name,
            company_name=company,
            email=email,
            phone=phone,
            country=country,
            is_active=True,
        )
        db.add(agent)
        db.flush()  # get agent.id

        # Assign to manager if provided
        if manager_name:
            manager = manager_map.get(manager_name.lower())
            if manager:
                db.add(models.ManagerAgentMapping(manager_id=manager.id, agent_id=agent.id))
            else:
                errors.append(f"Row {i}: Manager '{manager_name}' not found — agent created without manager.")

        created += 1

    db.commit()
    return schemas.BulkUploadResult(created=created, updated=updated, skipped=skipped, errors=errors)


# ─── Manager-Agent Mappings ────────────────────────────────────────────────────

@router.get("/manager/{manager_id}/agents", response_model=List[schemas.AgentOut])
def get_manager_agents(
    manager_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Get all agents assigned to a manager."""
    if current_user.role not in ("admin", "manager") and current_user.id != manager_id:
        raise HTTPException(status_code=403, detail="Forbidden")
    mappings = db.query(models.ManagerAgentMapping).filter(
        models.ManagerAgentMapping.manager_id == manager_id
    ).all()
    return [m.agent for m in mappings]


@router.post("/manager/{manager_id}/agents/{agent_id}")
def assign_agent_to_manager(
    manager_id: int,
    agent_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Assign an agent to a manager's responsibility."""
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Admin or manager required")
    existing = db.query(models.ManagerAgentMapping).filter_by(
        manager_id=manager_id, agent_id=agent_id
    ).first()
    if existing:
        return {"message": "Already assigned"}
    mapping = models.ManagerAgentMapping(manager_id=manager_id, agent_id=agent_id)
    db.add(mapping)
    db.commit()
    return {"message": "Agent assigned to manager"}


@router.delete("/manager/{manager_id}/agents/{agent_id}")
def unassign_agent_from_manager(
    manager_id: int,
    agent_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if current_user.role not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Admin or manager required")
    mapping = db.query(models.ManagerAgentMapping).filter_by(
        manager_id=manager_id, agent_id=agent_id
    ).first()
    if not mapping:
        raise HTTPException(status_code=404, detail="Mapping not found")
    db.delete(mapping)
    db.commit()
    return {"message": "Agent unassigned from manager"}


@router.get("/all-managers-with-agents")
def list_managers_with_agents(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Returns all manager-agent mappings (admin only)."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    managers = db.query(models.User).filter(
        models.User.role.in_(["admin", "manager"])
    ).all()
    result = []
    for m in managers:
        mappings = db.query(models.ManagerAgentMapping).filter_by(manager_id=m.id).all()
        result.append({
            "manager_id": m.id,
            "manager_name": m.full_name,
            "agents": [{"id": mp.agent.id, "name": mp.agent.name} for mp in mappings],
        })
    return result
